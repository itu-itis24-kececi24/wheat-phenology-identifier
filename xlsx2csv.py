#!/usr/bin/env python3

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def safe_filename(name: str) -> str:
    """Replace characters that are invalid or inconvenient in filenames."""
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", name).strip()
    return cleaned or "sheet"


def convert_xlsx_to_csv(
    xlsx_path: Path,
    output_dir: Path,
    sheet_name: str | None = None,
    delimiter: str = ",",
    encoding: str = "utf-8-sig",
) -> list[Path]:
    """
    Convert an XLSX workbook to CSV.

    If sheet_name is provided, only that sheet is converted.
    Otherwise, every sheet is converted into a separate CSV file.
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {xlsx_path}")

    if xlsx_path.suffix.lower() != ".xlsx":
        raise ValueError("The input file must have an .xlsx extension.")

    output_dir.mkdir(parents=True, exist_ok=True)

    # data_only=True exports calculated values instead of formula expressions,
    # provided that Excel has previously calculated and saved the workbook.
    workbook = load_workbook(
        filename=xlsx_path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise ValueError(
                    f"Sheet '{sheet_name}' was not found.\n"
                    f"Available sheets: {available}"
                )
            worksheets = [workbook[sheet_name]]
        else:
            worksheets = workbook.worksheets

        created_files: list[Path] = []

        for worksheet in worksheets:
            if len(worksheets) == 1:
                csv_filename = f"{xlsx_path.stem}.csv"
            else:
                csv_filename = (
                    f"{xlsx_path.stem}_{safe_filename(worksheet.title)}.csv"
                )

            csv_path = output_dir / csv_filename

            with csv_path.open(
                mode="w",
                newline="",
                encoding=encoding,
            ) as csv_file:
                writer = csv.writer(csv_file, delimiter=delimiter)

                for row in worksheet.iter_rows(values_only=True):
                    writer.writerow(
                        ["" if value is None else value for value in row]
                    )

            created_files.append(csv_path)

        return created_files

    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert an XLSX workbook into one or more CSV files."
    )

    parser.add_argument(
        "xlsx_file",
        type=Path,
        help="Path to the input .xlsx file.",
    )

    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=Path("."),
        help="Directory where CSV files will be created. Default: current directory",
    )

    parser.add_argument(
        "-s",
        "--sheet",
        help="Convert only the specified worksheet.",
    )

    parser.add_argument(
        "--delimiter",
        default=",",
        help=r"CSV delimiter. Examples: ',' or ';'. Default: ','",
    )

    args = parser.parse_args()

    if len(args.delimiter) != 1:
        parser.error("--delimiter must contain exactly one character.")

    try:
        created_files = convert_xlsx_to_csv(
            xlsx_path=args.xlsx_file,
            output_dir=args.output_dir,
            sheet_name=args.sheet,
            delimiter=args.delimiter,
        )

        for created_file in created_files:
            print(f"Created: {created_file.resolve()}")

        return 0

    except (FileNotFoundError, ValueError, PermissionError) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    except Exception as error:
        print(f"Unexpected error: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())